import pandas as pd
import openpyxl
import itertools
import json
import os
"""Merke: Ich muss noch sicher weiterlaufen lassen können, wenn ein Blatt fehlt (curve, rib, edge, surface)"""




def main():
    ### Anschlussanalyse ausführen
    # analyze_saf_connections('SAF_Analyser_Test')
    # analyze_saf_connections('21_22 L_TWP_Tragwerksmodell 0D Ausr Anschluss')              # erledigt
    # analyze_saf_connections('23_24 LTWP-V__Dachtragwerk 0D Ausr Anschluss')               # erledigt
    analyze_saf_connections('20220421MODEL REV01 0D Ausr Anschluss')


    ### GNN-Label-Generierung ausführen --> Erst nach annotierter Anschlussanalyse
    # generate_gnn_labels('SAF_Analyser_Test')
    # generate_gnn_labels('21_22 L_TWP_Tragwerksmodell 0D Ausr Anschluss')                  # erledigt
    # generate_gnn_labels('23_24 LTWP-V__Dachtragwerk 0D Ausr Anschluss')                   # erledigt
    return


def analyze_saf_connections(
        model_stem, export_excel=True, export_csv=True,
        start_index=0, end_index=None
        ):
    """
    Kurz Aufbau erklärt:
    "[ifc_stem] 0D Ausr" ==> Original-SAF von SCIA
    "[ifc_stem] 0D Ausr Anschluss" ==> Exportierte-SAF von Dlubal RFEM
    Auf die exportierte SAF diese Funktion anwenden, die bereits die Spalten in StructuralPointConnection einfügt
    Anschlussvarianten ggf. anpassen und bewerten.
    Dann kann zweite Funktion darauf aktiviert werden.
    """
    print(f"Lese Datei ein: {model_stem}...\n")

    # --- Schritt 0: Dateipfade vorbereiten ---
    # Basis-Pfad-Definition
    base_dir = os.path.dirname(__file__)
    saf_dir = os.path.join(base_dir, "Analysemodelle")

    # Dateinamen anpassen, letzte Bezeichnung ("0D Ausr Anschluss") kürzen
    cut_str = str("Ausr Anschluss")
    if model_stem.endswith(cut_str):
        # Entfernt letzte Zeichen + 3 Zeichen davor (z.B. " 0D")
        cut_length = len(cut_str) + 3
        ifc_name = model_stem[:-cut_length].strip()
    else:
        ifc_name = model_stem.strip()

    # Dateipfade zusammenbauen
    path_saf = os.path.join(saf_dir, f"{model_stem}.xlsx")
    output_csv = os.path.join(saf_dir, f"{ifc_name} Anschlussanalyse.csv")
    # ----------------------------------------------------


    # --- Schritt 1: Mapping ---
    #  Einlesen der definierten Tabellenblätter
    try:
        xls = pd.ExcelFile(path_saf)
    except Exception as e:
        print(f"Kritischer Fehler beim Öffnen der Excel-Datei: {e}")
        return
    
    sheet_names = xls.sheet_names

    # Standardmäßig erzeugte Blätter prüfen
    required_sheets = ['StructuralPointConnection', 'StructuralCurveMember', 'StructuralSurfaceMember']
    for sheet in required_sheets:
        if sheet not in sheet_names:
            print(f"KRITISCHER FEHLER: Das zwingend erforderliche Blatt '{sheet}' fehlt in der Datei.")
            return
    # Zwingende Blätter einlesen
    # Als Input geht sowohl xls als auch path_saf --> Zur Demonstration habe ich beides drin
    df_nodes = pd.read_excel(xls, sheet_name='StructuralPointConnection')
    df_curve = pd.read_excel(xls, sheet_name='StructuralCurveMember')
    df_surface = pd.read_excel(xls, sheet_name='StructuralSurfaceMember')

    # Optionale Blätter einlesen (mit Fallback auf leere Blätter)
    if 'StructuralCurveMemberRib' in sheet_names:
        df_rib = pd.read_excel(path_saf, sheet_name='StructuralCurveMemberRib')
    else:
        print(f"[INFO] Tabellenblatt 'StructuralCurveMemberRib' fehlt. Rippen werden ignoriert.")
        df_rib = pd.DataFrame()    
    if 'StructuralCurveEdge' in sheet_names:
        df_edge = pd.read_excel(path_saf, sheet_name='StructuralCurveEdge')
    else:
        print(f"[INFO] Tabellenblatt 'StructuralCurveEdge' fehlt. Kanten werden ignoriert.")
        df_edge = pd.DataFrame()

    # Spaltennamen bereinigen
    for df in [df_nodes, df_curve, df_rib, df_surface, df_edge]:
        if not df.empty:
            df.columns = df.columns.str.strip()

    # Feste Reihenfolge für Namensgebung
    type_order = ['Stütze', 'Balken', 'Verband', 'Platte', 'Wand']
    
    # Mapping von SAF-Typen (Englisch) zu Output-Typen (Deutsch)
    type_mapping = {
        'Column': 'Stütze',
        'Beam': 'Balken',
        'Member': 'Verband',
        'Slab': 'Platte',
        'Wall': 'Wand'
    }
    # ----------------------------------------------------


    # --- Schritt 2: Verbindungsanalyse ---
    # Dictionary zur Speicherung initialisieren (Schlüssel = Knoten-Name)
    # Dictionary speichert pro Knoten weiteres Dictionary: {Elementname: Elementtyp}
    # Dies verhindert doppelte Zählungen desselben Elements am selben Knoten
    node_connections = {str(node): {} for node in df_nodes['Name']}
    missing_nodes = set()
    unrecognized_types = set()

    # Hilfsfunktion zum Zuweisen der Bauteile zu den jeweiligen Knoten
    def add_connections(element_name, nodes_string, saf_type):
        if pd.notna(nodes_string) and pd.notna(saf_type):
            mapped_type = type_mapping.get(str(saf_type).strip())
            
            if not mapped_type:
                # Unbekannter Typ, wird hier gesammelt und später ausgegeben
                unrecognized_types.add(str(saf_type))
                return 

            nodes = str(nodes_string).split(';') # SPlit erzeugt Liste
            for node in nodes:
                node = node.strip()
                if node in node_connections:
                    # Speichert/Überschreibt das Element mit seinem Typ
                    node_connections[node][str(element_name)] = mapped_type
                elif node:
                    # Wenn der Knoten nicht leer ist, aber im Dictionary fehlt, wird er hier gesammelt
                    missing_nodes.add(node)

    # Lookup-Dictionary für Flächenelemente (Name -> Typ)
    # Erforderlich, da in StructuralCurveEdge der Typ nicht mehr steht
    """Hinweis für folgenden Code:
    - .iterrows gibt immer 2 Werte zurück (Zeilenindex und Zellenwert)
    - Deshalb brauche ich _ als Wegwerf-Variable
    - Alte Variante row['Type'] == row.get('Type') 
    - get gibt None zurück, wenn Spalte fehlt, statt Fehler zu werfen"""

    surface_types = {}
    for _, row in df_surface.iterrows():
        if pd.notna(row.get('Name')) and pd.notna(row.get('Type')):
            surface_types[str(row.get('Name'))] = str(row.get('Type'))

    # 1. Stabelemente (StructuralCurveMember)
    for _, row in df_curve.iterrows():
        add_connections(row.get('Name'), row.get('Nodes'), row.get('Type'))

    # 2. Rippen (StructuralCurveMemberRib))
    # Schleife überspringt sich selbst, wenn df leer ist
    for _, row in df_rib.iterrows():
        add_connections(row.get('Name'), row.get('Nodes'), 'Beam')

    # 3. Flächenelemente (StructuralSurfaceMember)
    for _, row in df_surface.iterrows():
        element_name = row.get('Name')
        saf_type = row.get('Type')
        # Normale Ranknoten verarbeiten
        if pd.notna(row.get('Nodes')):
            add_connections(element_name, row.get('Nodes'), saf_type)
        # Interne Knoten verarbeiten
        if pd.notna(row.get('Internal nodes')):
            add_connections(element_name, row.get('Internal nodes'), saf_type)

    # 4. Kanten von Flächenelementen (StructuralCurveEdge)
    # Schleife überspringt sich selbst, wenn df leer ist
    for _, row in df_edge.iterrows():
        element_name = row.get('2D member')
        nodes_string = row.get('Nodes')
        
        if pd.notna(element_name) and pd.notna(nodes_string):
            # Typ aus dem Lookup-Dictionary abrufen
            saf_type = surface_types.get(str(element_name))
            if saf_type:
                add_connections(element_name, nodes_string, saf_type)


    # 5. Dynamische Printausgabe der gewählten Range im Titel durch Slicing
    print(f"""\n=== Verbindungsanalyse der Knoten (Index {start_index} bis {end_index if end_index else 'Ende'}) ===""")
    # Dictionary zum Zwischenspeichern der Ergebnisse für den Excel-Export initialisieren
    calculated_results = {}
    
    # Liste für die CSV-Ausgabe initialisieren
    export_data = []
    
    for node_id, elements_dict in list(node_connections.items())[start_index:end_index]:
        total_elements = len(elements_dict)
        
        # Logik für Bezeichnung der Anschlussvariante (Kategoriename)
        if total_elements <= 1:
            category_name = "..."
        else:
            # Zählt die Vorkommen der Typen für diesen Knoten
            type_counts = {t: 0 for t in type_order}
            for elem_type in elements_dict.values():
                type_counts[elem_type] += 1
            
            # Baut den String in der strikten Reihenfolge auf
            name_parts = []
            for t in type_order:
                if type_counts[t] > 0:
                    name_parts.append(f"{t}{type_counts[t]}")
            
            category_name = "_".join(name_parts)

        element_names = list(elements_dict.keys())
        elements_string = ', '.join(element_names) if element_names else 'Keine'

        # Konsolenausgabe - detailliert
        # print(f"Knoten {node_id}:")
        # print(f"  - Kategorie: {category_name}")
        # print(f"  - Beteiligte Elemente ({total_elements}): {elements_string}")
        # print("-" * 30)

        # Konsolenausgabe - kompakt
        # print(f"Knoten {node_id}: {category_name}")
        # print(f"{category_name}")

        # Daten für die spätere Excel-Ausgabe zwischenspeichern
        calculated_results[node_id] = {
            'Anschlussvariante': category_name,
            'Anzahl': total_elements,
            'Elemente': elements_string
        }
        
        # Datenzeile für CSV-Format anfügen
        export_data.append({
            'Knoten': node_id,
            'Anschlussvariante': category_name,
            'Anzahl': total_elements,
            'Elemente': elements_string
        })
    # ----------------------------------------------------

    
    # --- Schritt 3: Export der Ergebnisse ---
    # Ergebnisse in bestehende Excel-Datei schreiben
    print("\nSchreibe Ergebnisse in die Excel-Datei zurück...")
    if export_excel:
        try:
            # Arbeitsmappe laden
            wb = openpyxl.load_workbook(path_saf)

            # Tabellenblatt dynamisch finden
            sheet_name = 'StructuralPointConnection'
            if sheet_name not in wb.sheetnames:
                print(f"FEHLER: Tabellenblatt '{sheet_name}' in der Excel-Datei nicht gefunden.")
                return
        
            ws = wb[sheet_name]

            # Suchen der Spalte "Name" und der ersten leeren Spalte
            name_col_idx = None
            max_col_idx = ws.max_column

            for col in range(1, max_col_idx + 1):
                cell_value = ws.cell(row=1, column=col).value
                if cell_value and str(cell_value).strip() == 'Name':
                    name_col_idx = col
                    break
            
            if not name_col_idx:
                print("FEHLER: Spalte 'Name' nicht gefunden.")
                return
            
            # Indizes für die neuen Spalten bestimmen
            label_col_idx = max_col_idx + 1
            cat_col_idx = max_col_idx + 2
            count_col_idx = max_col_idx + 3
            elem_col_idx = max_col_idx + 4
            ergänzt_col_idx = max_col_idx + 5
            entfernt_col_idx = max_col_idx + 6
            kommentar_col_idx = max_col_idx + 7

            # Header schreiben
            ws.cell(row=1, column=label_col_idx, value='Label')
            ws.cell(row=1, column=cat_col_idx, value='Anschlussvariante')
            ws.cell(row=1, column=count_col_idx, value='Anzahl')
            ws.cell(row=1, column=elem_col_idx, value='Elemente')
            ws.cell(row=1, column=ergänzt_col_idx, value='fehlend')
            ws.cell(row=1, column=entfernt_col_idx, value='zuviel')
            ws.cell(row=1, column=kommentar_col_idx, value='Kommentar')

            # Zeile für Zeile durchlaufen und Werte aus calculated_results eintragen
            for row in range(2, ws.max_row + 1):
                node_name = ws.cell(row=row, column=name_col_idx).value
                if node_name and str(node_name).strip() in calculated_results:
                    result = calculated_results[str(node_name)]
                    ws.cell(row=row, column=cat_col_idx, value=result['Anschlussvariante'])
                    ws.cell(row=row, column=count_col_idx, value=result['Anzahl'])
                    ws.cell(row=row, column=elem_col_idx, value=result['Elemente'])
            
            # Datei abspeichern
            wb.save(path_saf)
            print(f">>> ERFOLG: Ergebnisse wurden in '{path_saf}' zurückgeschrieben.")

        except PermissionError:
            print(f"\n!!! KRITISCHER FEHLER: Die Datei '{path_saf}' ist in Excel geöffnet! Bitte schließe die Datei und starte das Skript erneut. !!!")
        except Exception as e:
            print(f"\n!!! FEHLER beim Schreiben in die Excel-Datei: {e} !!!")

    # Ergbnisse als DataFrame aus der Liste erstellen und als CSV speichern
    if export_csv:
        print("\nSchreibe Ergebnisse in CSV-Datei...")
        if export_data:        
            df_export = pd.DataFrame(export_data)
            df_export.to_csv(output_csv, index=False, 
                                sep=';', encoding='utf-8-sig')
            print(f">>> ERFOLG: Die Ergebnisse wurden in '{output_csv}' gespeichert. <<<")

    # Ausgabe der Warnung am Ende des Skripts
    if unrecognized_types:
        print("\n!!! ANALYSE: Unbekannte Bauteiltypen gefunden !!!")
        print("Folgende Werte konnten nicht zugeordnet werden:")
        print(", ".join(unrecognized_types))
        
    if missing_nodes:
        print("\n!!! WARNUNG: Inkonsistente Knotenreferenzen !!!")
        print(", ".join(sorted(missing_nodes)))





"""----- ZWEITER TEIL ----- ZWEITER TEIL ----- ZWEITER TEIL -----"""





def generate_gnn_labels(model_stem, export_csv=True, export_json=True):
    print(f"Lese Datei ein: {model_stem}...\n")
    # --- Schritt 0: Dateipfade vorbereiten ---
    # Basis-Pfad-Definition
    base_dir = os.path.dirname(__file__)
    saf_dir = os.path.join(base_dir, "Analysemodelle")
    dir_2_3 = os.path.join(base_dir, "2_3 Edge_Labeling")

    os.makedirs(dir_2_3, exist_ok=True) # Zielordner erstellen. falls er nicht existiert

    # Dateinamen anpassen, letzte Bezeichnung ("0D Ausr Anschluss") kürzen
    cut_str = str("Ausr Anschluss")
    if model_stem.endswith(cut_str):
        # Entfernt letzte Zeichen + 3 Zeichen davor (z.B. " 0D")
        cut_length = len(cut_str) + 3
        ifc_name = model_stem[:-cut_length].strip()
    else:
        ifc_name = model_stem.strip()

    # Dateipfade zusammenbauen
    path_saf = os.path.join(saf_dir, f"{model_stem}.xlsx")
    output_json = os.path.join(dir_2_3, f"{ifc_name} Labels.json")
    output_csv = os.path.join(dir_2_3, f"{ifc_name} Labels.csv")
    # ----------------------------------------------------


    # --- Schritt 1: Mapping und Total Labels generieren ---
    try:
        df_nodes = pd.read_excel(path_saf, sheet_name='StructuralPointConnection')
        df_curve = pd.read_excel(path_saf, sheet_name='StructuralCurveMember')
        df_rib = pd.read_excel(path_saf, sheet_name='StructuralCurveMemberRib')
        df_surface = pd.read_excel(path_saf, sheet_name='StructuralSurfaceMember')
    except Exception as e:
        print(f"Fehler beim einlesen: {e}")
        return
    
    # Spalten bereinigen
    for df in [df_nodes, df_curve, df_rib, df_surface]:
        df.columns = df.columns.str.strip()

    saf_to_ifc = {}
    all_ifc_guids = set()

    # SAF und IFC-Elemente mappen
    for df in [df_curve, df_rib, df_surface]:
        # Prüfen ob GUID existiert
        if 'GUID' not in df.columns:
            continue

        for _, row in df.iterrows():
            saf_name = str(row.get('Name')).strip()
            ifc_guid = str(row.get('GUID')).strip()

            if saf_name != 'nan' and ifc_guid != 'nan':
                saf_to_ifc[saf_name] = ifc_guid
                all_ifc_guids.add(ifc_guid)
    
    print(f"Gefundene einzigartige IFC-Bauteile: {len(all_ifc_guids)}")

    # Total Labels: Alle möglichen Kombinationen im ganzen Bauwerk (ungerichtet)
    Total_Labels = set(tuple(sorted(comb)) for comb in itertools.combinations(all_ifc_guids, 2))
    print(f"Generierte Total_Labels (Alle möglichen Kanten): {len(Total_Labels)}")
    # ----------------------------------------------------


    # --- SCHRITT 2: GroundTruth (GT) und Rule_Labels ---
    GT_Labels = set()
    Rule_Labels = set()

    # Hilfskunktion zum Verarbeiten von Strings wie "CM25-SFC46, CM25-SFC47"
    def parse_edges(edge_string):
        parsed_edges = set()
        if pd.isna(edge_string) or str(edge_string).strip() == '':
            return parsed_edges
        
        parts = str(edge_string).split(',')
        for part in parts:
            if '-' in part:
                elements = part.split('-', 1)
                if len(elements) == 2:
                    saf_a, saf_b = elements[0].strip(), elements[1].strip()
                    # Nur hinzufügen, wenn beide SAF-Namen eine GUID haben
                    if saf_a in saf_to_ifc and saf_b in saf_to_ifc:
                        ifc_a = saf_to_ifc[saf_a]
                        ifc_b = saf_to_ifc[saf_b]
                        edge_tuple = tuple(sorted((ifc_a, ifc_b)))
                        parsed_edges.add(edge_tuple)
        return parsed_edges
        
    
    for _, row in df_nodes.iterrows():
        label = str(row.get('Label')).strip().lower()
        if label not in ['x', 'o']:
            continue

        elemente_str = str(row.get('Elemente'))

        # GT_Labels für diesen Knoten generieren
        saf_elements_at_node = [e.strip() for e in elemente_str.split(',')]
        ifc_elements_at_node = set()

        for saf in saf_elements_at_node:
            if saf in saf_to_ifc:
                ifc_elements_at_node.add(saf_to_ifc[saf])

        # Alle Kombinationen der beteiligten Elemente am Knoten = Ground Truth
        node_gt_edges = set(tuple(sorted(comb)) for comb in itertools.combinations(ifc_elements_at_node, 2))
        GT_Labels.update(node_gt_edges)

        # Modus-Logik für Rule_Labels
        node_rule_edges = set(node_gt_edges)  # Startet mit GT-Labels als perfekte Basis

        fehlend_str = str(row.get('fehlend')).strip().lower()
        zuviel_str = str(row.get('zuviel')).strip().lower()

        # 1. Fehlende Verbindungen abziehen
        if fehlend_str == 'alle':
            # Konverter hat hier keine Elemente miteinander verbunden, trotz Vollkopplung.
            node_rule_edges.clear()  
        elif fehlend_str != 'nan':
            edges_to_remove = parse_edges(row.get('fehlend'))
            node_rule_edges.difference_update(edges_to_remove)
        
        # 2. Übermäßige Verbindungen hinzufügen
        if zuviel_str != 'nan':
            edges_to_add = parse_edges(row.get('zuviel'))
            node_rule_edges.update(edges_to_add)
        
        Rule_Labels.update(node_rule_edges)
    
    print(f"Generierte GT_Labels: {len(GT_Labels)}")
    print(f"Generierte Rule_Labels: {len(Rule_Labels)}")
    # ----------------------------------------------------


    # --- SCHRITT 3: Metriken berechnen ---
    tp = len(GT_Labels.intersection(Rule_Labels)) # Prüft die Schnittmenge beider Labels
    fp = len(Rule_Labels.difference(GT_Labels)) # Realität sagt "ja", Konverter hat sie übersehen
    fn = len(GT_Labels.difference(Rule_Labels)) # Realität sagt "nein", Konverter hat sie fälschlicherweise verbunden
    tn = len(Total_Labels.difference(GT_Labels.union(Rule_Labels))) # Keine Verbindung in Realität und Konverter

    total_possible = len(Total_Labels)
    accuracy = (tp + tn) / total_possible if total_possible > 0 else 0
    precision = tp / (tp + fp) if (tp + fp) > 0 else 0
    recall = tp / (tp + fn) if (tp + fn) > 0 else 0
    f1_score = 2 * (precision * recall) / (precision + recall) if (precision + recall) > 0 else 0

    print("\n=== AUSWERTUNG DER KONVERTER-REGELN (Metriken) ===")
    print(f"True Positives (TP):  {tp:<6} (Richtig erkannt)")
    print(f"True Negatives (TN):  {tn:<6} (Richtig ignoriert)")
    print(f"False Positives (FP): {fp:<6} (Zu viel verbunden)")
    print(f"False Negatives (FN): {fn:<6} (Vergessen)")
    print("-" * 50)
    print(f"Accuracy:  {accuracy:.4f} ({(accuracy*100):.2f} %)")
    print(f"Precision: {precision:.4f} ({(precision*100):.2f} %)")
    print(f"Recall:    {recall:.4f} ({(recall*100):.2f} %)")
    print(f"F1-Score:  {f1_score:.4f}")
    print("==================================================")
    # ----------------------------------------------------


    # --- SCHRITT 4: Ergebnisse dynamisch exportieren ---
    # Konvertierung in strukturierte Dictionaries
    def set_to_dict_list(edge_set, label_type):
        return [{'GUID_A': edge[0], 'GUID_B': edge[1], 'Label_Type': label_type} for edge in edge_set]
    
    export_data = set_to_dict_list(Total_Labels, 'Total') + \
                set_to_dict_list(GT_Labels, 'GT') + \
                set_to_dict_list(Rule_Labels, 'Rule')
    
    df_export = pd.DataFrame(export_data)
    
    # JSON-Export
    if export_json:
        with open(output_json, 'w', encoding='utf-8') as f:
            json.dump(export_data, f, indent=4)
    
    # CSV Export
    if export_csv:
        df_export.to_csv(output_csv, index=False, sep=';', encoding='utf-8-sig')

    print(f"\n>>> ERFOLG: GNN-Labels wurden in '{output_csv}' und '{output_json}' gespeichert. <<<")
    print(f"\n--- ___KOPIERE DEN TEXT IN DIE CSV !!___ ---")
    # ----------------------------------------------------





"""Main-Part"""
if __name__ == "__main__":
    main()
