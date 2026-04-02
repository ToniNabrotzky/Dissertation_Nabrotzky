import itertools
import json
import math
import openpyxl
import os
import pandas as pd
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter


def main():
    ### Anschlussanalyse ausführen
    analyze_saf_connections('SAF_Analyser_Test - Kopie', check_internal_nodes=True)
    # analyze_saf_connections('21_22 L_TWP_Tragwerksmodell 0D Ausr Anschluss')              # erledigt
    # analyze_saf_connections('23_24 LTWP-V__Dachtragwerk 0D Ausr Anschluss')               # erledigt
    # analyze_saf_connections('20220421MODEL REV01 0D Ausr Anschluss')                      # erledigt
    # analyze_saf_connections('202102183458-Model 0d Ausr Anschluss')                       # erledigt
    # analyze_saf_connections('Grethes-hus-bok-2 0d Ausr Anschluss')                        # erledigt
    # analyze_saf_connections('Vectorworks2016-IFC2x3-EQUA_IDA_ICE 0d Ausr Anschluss')      # erledigt


    ### GNN-Label-Generierung ausführen --> Erst nach annotierter Anschlussanalyse
    # generate_gnn_labels('SAF_Analyser_Test - Kopie')
    # generate_gnn_labels('21_22 L_TWP_Tragwerksmodell 0D Ausr Anschluss')                  # erledigt
    # generate_gnn_labels('23_24 LTWP-V__Dachtragwerk 0D Ausr Anschluss')                   # erledigt
    # generate_gnn_labels('20220421MODEL REV01 0D Ausr Anschluss')                          # erledigt
    # generate_gnn_labels('202102183458-Model 0d Ausr Anschluss')                           # erledigt
    # generate_gnn_labels('Grethes-hus-bok-2 0d Ausr Anschluss')                            # erledigt
    return


# --- Geometrische Hilfsfunktionen für die Prüfung interner Knoten ---
def _distance_3d(p1, p2):
    return math.sqrt((p1[0]-p2[0])**2 + (p1[1]-p2[1])**2 + (p1[2]-p2[2])**2)

def _is_point_on_segment(p, a, b, tol=1e-3):
    """Prüft, ob Punkt p auf der Strecke ab liegt (mit Toleranz in Metern)."""
    d_ab = _distance_3d(a, b)
    if d_ab < 1e-6: return False
    d_ap = _distance_3d(a, p)
    d_pb = _distance_3d(p, b)
    return abs(d_ap + d_pb - d_ab) < tol

def _is_point_in_polygon_3d(p, poly_points, tol=1e-3):
    """Prüft, ob Punkt p innerhalb eines 3D-Polygons oder auf dessen Rand liegt."""
    if len(poly_points) < 3: return False
    p0 = poly_points[0]
    p1 = poly_points[1]
    normal = None
    # Finde die Ebene des Polygons
    for i in range(2, len(poly_points)):
        p2 = poly_points[i]
        v1 = (p1[0]-p0[0], p1[1]-p0[1], p1[2]-p0[2])
        v2 = (p2[0]-p0[0], p2[1]-p0[1], p2[2]-p0[2])
        nx = v1[1]*v2[2] - v1[2]*v2[1]
        ny = v1[2]*v2[0] - v1[0]*v2[2]
        nz = v1[0]*v2[1] - v1[1]*v2[0]
        length = math.sqrt(nx*nx + ny*ny + nz*nz)
        if length > 1e-6:
            normal = (nx/length, ny/length, nz/length)
            break
    
    # Fallback, falls alle Punkte kollinear sind
    if not normal:
        for i in range(len(poly_points)-1):
            if _is_point_on_segment(p, poly_points[i], poly_points[i+1], tol): return True
        return _is_point_on_segment(p, poly_points[-1], poly_points[0], tol)
        
    # Ist der Punkt auf derselben Ebene?
    v_p = (p[0]-p0[0], p[1]-p0[1], p[2]-p0[2])
    if abs(v_p[0]*normal[0] + v_p[1]*normal[1] + v_p[2]*normal[2]) > tol: 
        return False
        
    # Auf dominante Achse projizieren (2D Ray-Casting)
    nx_abs, ny_abs, nz_abs = abs(normal[0]), abs(normal[1]), abs(normal[2])
    if nz_abs >= nx_abs and nz_abs >= ny_abs: p_2d, poly_2d = (p[0], p[1]), [(pt[0], pt[1]) for pt in poly_points]
    elif ny_abs >= nx_abs and ny_abs >= nz_abs: p_2d, poly_2d = (p[0], p[2]), [(pt[0], pt[2]) for pt in poly_points]
    else: p_2d, poly_2d = (p[1], p[2]), [(pt[1], pt[2]) for pt in poly_points]
        
    x, y = p_2d
    n = len(poly_2d)
    inside = False
    p1x, p1y = poly_2d[0]
    for i in range(n + 1):
        p2x, p2y = poly_2d[i % n]
        if math.hypot(p2x - p1x, p2y - p1y) > 1e-6 and abs(math.hypot(x - p1x, y - p1y) + math.hypot(p2x - x, p2y - y) - math.hypot(p2x - p1x, p2y - p1y)) < tol: return True
        if min(p1y, p2y) < y <= max(p1y, p2y) and x <= max(p1x, p2x):
            if p1y != p2y and (p1x == p2x or x <= (y - p1y) * (p2x - p1x) / (p2y - p1y) + p1x): inside = not inside
        p1x, p1y = p2x, p2y
    return inside
# ----------------------------------------------------------------------


def analyze_saf_connections(
        model_stem, export_excel=True, export_csv=False,
        start_index=0, end_index=None, check_types=True,
        check_internal_nodes=True
        ):
    """
    Kurz Aufbau erklärt:
    "[ifc_stem] 0D Ausr" ==> Original-SAF von SCIA
    "[ifc_stem] 0D Ausr Anschluss" ==> Exportierte-SAF von Dlubal RFEM
    Auf die exportierte SAF diese Funktion anwenden, die bereits die Spalten in StructuralPointConnection einfügt
    Anschlussvarianten ggf. anpassen und bewerten.
    Dann kann zweite Funktion darauf aktiviert werden.
    """
    print(f"Lese Datei ein: {model_stem}...\n")

    # --- Schritt 0: Dateipfade vorbereiten ---
    # Basis-Pfad-Definition
    base_dir = os.path.dirname(__file__)
    saf_dir = os.path.join(base_dir, "Analysemodelle")

    # Dateinamen anpassen, letzte Bezeichnung ("0D Ausr Anschluss") kürzen
    cut_str = str("Ausr Anschluss")
    if model_stem.endswith(cut_str):
        # Entfernt letzte Zeichen + 3 Zeichen davor (z.B. " 0D")
        cut_length = len(cut_str) + 3
        ifc_name = model_stem[:-cut_length].strip()
    else:
        ifc_name = model_stem.strip()

    # Dateipfade zusammenbauen
    path_saf = os.path.join(saf_dir, f"{model_stem}.xlsx")
    output_csv = os.path.join(saf_dir, f"{ifc_name} Anschlussanalyse.csv")

    # Sicherheitsabfrage zur Prüfung der SAF-Typen
    if check_types:
        print(f"\n---SICHERHEITSABFRAGE---")
        while True:
            user_confirm = input(f"Hast du alle SAF-Bauteiltypen in '{model_stem}' geprüft und korrigiert? (j/n):").strip().lower()
            if user_confirm in ["j", "ja", 'y', 'yes', '']:
                break
            elif user_confirm in ['n', 'nein', 'no']:
                print("\n>>> Skript abgebrochen. Bitte passe die Typen in der Excel-Datei zuerst an und starte das Skript danach neu. <<<")
                return
            else:
                print("Ungültige Eingabe. Bitte antworte mit 'j' für Ja oder 'n' für Nein.")
        print("--------------------------\n")



    # ----------------------------------------------------


    # --- Schritt 1: Mapping ---
    #  Einlesen der definierten Tabellenblätter
    try:
        xls = pd.ExcelFile(path_saf)
    except Exception as e:
        print(f"Kritischer Fehler beim Öffnen der Excel-Datei: {e}")
        return
    
    sheet_names = xls.sheet_names

    # Standardmäßig erzeugte Blätter prüfen
    required_sheets = ['StructuralPointConnection', 'StructuralCurveMember', 'StructuralSurfaceMember']
    for sheet in required_sheets:
        if sheet not in sheet_names:
            print(f"KRITISCHER FEHLER: Das zwingend erforderliche Blatt '{sheet}' fehlt in der Datei.")
            return
    # Zwingende Blätter einlesen
    # Als Input geht sowohl xls als auch path_saf --> Zur Demonstration habe ich beides drin
    df_nodes = pd.read_excel(xls, sheet_name='StructuralPointConnection')
    df_curve = pd.read_excel(xls, sheet_name='StructuralCurveMember')
    df_surface = pd.read_excel(xls, sheet_name='StructuralSurfaceMember')

    # Optionale Blätter einlesen (mit Fallback auf leere Blätter)
    if 'StructuralCurveMemberRib' in sheet_names:
        df_rib = pd.read_excel(path_saf, sheet_name='StructuralCurveMemberRib')
    else:
        print(f"[INFO] Tabellenblatt 'StructuralCurveMemberRib' fehlt. Rippen werden ignoriert.")
        df_rib = pd.DataFrame(columns=['Name', 'Type', 'Id', '2D member', 'Nodes'])    
    if 'StructuralCurveEdge' in sheet_names:
        df_edge = pd.read_excel(path_saf, sheet_name='StructuralCurveEdge')
    else:
        print(f"[INFO] Tabellenblatt 'StructuralCurveEdge' fehlt. Kanten werden ignoriert.")
        df_edge = pd.DataFrame(columns=['Name', 'Type', 'Id', '2D member', 'Nodes'])

    # Spaltennamen bereinigen
    for df in [df_nodes, df_curve, df_rib, df_surface, df_edge]:
        if not df.empty:
            df.columns = df.columns.str.strip()

    # Feste Reihenfolge für Namensgebung
    type_order = ['Stütze', 'Balken', 'Verband', 'Platte', 'Wand']
    
    # Mapping von SAF-Typen (Englisch) zu Output-Typen (Deutsch)
    type_mapping = {
        'Beam': 'Balken',
        'Column': 'Stütze',
        'Member': 'Verband',
        'Slab': 'Platte',
        'Plate': 'Platte',
        'Wall': 'Wand'
    }
    # ----------------------------------------------------


    # --- Schritt 2: Verbindungsanalyse ---
    # Dictionary zur Speicherung initialisieren (Schlüssel = Knoten-Name)
    # Dictionary speichert pro Knoten weiteres Dictionary: {Elementname: Elementtyp}
    # Dies verhindert doppelte Zählungen desselben Elements am selben Knoten
    node_connections = {str(node): {} for node in df_nodes['Name']}
    missing_nodes = set()
    unrecognized_types = set()

    # Hilfsfunktion zum Zuweisen der Bauteile zu den jeweiligen Knoten
    def add_connections(element_name, nodes_string, saf_type):
        if pd.notna(nodes_string) and pd.notna(saf_type):
            mapped_type = type_mapping.get(str(saf_type).strip())
            
            if not mapped_type:
                # Unbekannter Typ, wird hier gesammelt und später ausgegeben
                unrecognized_types.add(str(saf_type))
                return 

            nodes = str(nodes_string).split(';') # Split erzeugt Liste
            for node in nodes:
                node = node.strip()
                if node in node_connections:
                    # Speichert/Überschreibt das Element mit seinem Typ
                    node_connections[node][str(element_name)] = mapped_type
                elif node:
                    # Wenn der Knoten nicht leer ist, aber im Dictionary fehlt, wird er hier gesammelt
                    missing_nodes.add(node)
                    
    # Koordinaten der Knoten extrahieren, falls geometrische Analyse aktiv ist
    points_dict = {}
    if check_internal_nodes and not df_nodes.empty:
        x_col = next((c for c in df_nodes.columns if c in ['X', 'x', 'Coord X', 'Coordinate X', 'Coordinate X [m]']), None)
        y_col = next((c for c in df_nodes.columns if c in ['Y', 'y', 'Coord Y', 'Coordinate Y', 'Coordinate Y [m]']), None)
        z_col = next((c for c in df_nodes.columns if c in ['Z', 'z', 'Coord Z', 'Coordinate Z', 'Coordinate Z [m]']), None)
        if x_col and y_col and z_col:
            for _, row in df_nodes.iterrows():
                if pd.notna(row.get('Name')):
                    points_dict[str(row.get('Name')).strip()] = (float(row[x_col]), float(row[y_col]), float(row[z_col]))
        else:
            print("[WARNUNG] Koordinatenspalten (X, Y, Z) in 'StructuralPoint' nicht gefunden. 'check_internal_nodes' wird deaktiviert.")
            check_internal_nodes = False

    # Lookup-Dictionary für Flächenelemente (Name -> Typ)
    # Erforderlich, da in StructuralCurveEdge der Typ nicht mehr steht
    """Hinweis für folgenden Code:
    - .iterrows gibt immer 2 Werte zurück (Zeilenindex und Zellenwert)
    - Deshalb brauche ich _ als Wegwerf-Variable
    - Alte Variante row['Type'] == row.get('Type') 
    - get gibt None zurück, wenn Spalte fehlt, statt Fehler zu werfen"""

    surface_types = {}
    for _, row in df_surface.iterrows():
        if pd.notna(row.get('Name')) and pd.notna(row.get('Type')):
            surface_types[str(row.get('Name'))] = str(row.get('Type'))

    # 1. Stabelemente (StructuralCurveMember)
    for _, row in df_curve.iterrows():
        add_connections(row.get('Name'), row.get('Nodes'), row.get('Type'))

    # 2. Rippen (StructuralCurveMemberRib)
    # Schleife überspringt sich selbst, wenn df leer ist
    for _, row in df_rib.iterrows():
        add_connections(row.get('Name'), row.get('Nodes'), 'Beam')

    # 3. Flächenelemente (StructuralSurfaceMember)
    for _, row in df_surface.iterrows():
        element_name = row.get('Name')
        saf_type = row.get('Type')
        # Normale Ranknoten verarbeiten
        if pd.notna(row.get('Nodes')):
            add_connections(element_name, row.get('Nodes'), saf_type)
        # Interne Knoten verarbeiten
        if pd.notna(row.get('Internal nodes')):
            add_connections(element_name, row.get('Internal nodes'), saf_type)

    # 4. Kanten von Flächenelementen (StructuralCurveEdge)
    # Schleife überspringt sich selbst, wenn df leer ist
    for _, row in df_edge.iterrows():
        element_name = row.get('2D member')
        nodes_string = row.get('Nodes')
        
        if pd.notna(element_name) and pd.notna(nodes_string):
            # Typ aus dem Lookup-Dictionary abrufen
            saf_type = surface_types.get(str(element_name))
            if saf_type:
                add_connections(element_name, nodes_string, saf_type)

    # --- Automatische Überprüfung auf interne Knoten (Geometrisch) ---
    if check_internal_nodes and points_dict:
        print("[INFO] Führe geometrische Prüfung auf interne Knoten (Linien/Flächen) durch...")
        for node_id in node_connections.keys():
            if str(node_id) not in points_dict:
                continue
            p = points_dict[str(node_id)]
            
            # Prüfung gegen 1D-Elemente (Liniensegmente)
            for _, row in df_curve.iterrows():
                element_name = str(row.get('Name')).strip()
                if element_name == 'nan' or element_name in node_connections[node_id]: continue
                
                nodes_string = row.get('Nodes')
                if pd.notna(nodes_string):
                    poly_points = [points_dict[n.strip()] for n in str(nodes_string).split(';') if n.strip() in points_dict]
                    if len(poly_points) >= 2:
                        for i in range(len(poly_points) - 1):
                            if _is_point_on_segment(p, poly_points[i], poly_points[i+1]):
                                mapped_type = type_mapping.get(str(row.get('Type')).strip())
                                if mapped_type: node_connections[node_id][element_name] = mapped_type
                                break

            # Prüfung gegen 2D-Elemente (Polygone)
            for _, row in df_surface.iterrows():
                element_name = str(row.get('Name')).strip()
                if element_name == 'nan' or element_name in node_connections[node_id]: continue
                
                nodes_string = row.get('Nodes')
                if pd.notna(nodes_string):
                    poly_points = [points_dict[n.strip()] for n in str(nodes_string).split(';') if n.strip() in points_dict]
                    if len(poly_points) >= 3:
                        if _is_point_in_polygon_3d(p, poly_points):
                            mapped_type = type_mapping.get(str(row.get('Type')).strip())
                            if mapped_type: node_connections[node_id][element_name] = mapped_type

    # 5. Dynamische Printausgabe der gewählten Range im Titel durch Slicing
    print(f"""\n=== Verbindungsanalyse der Knoten (Index {start_index} bis {end_index if end_index else 'Ende'}) ===""")
    # Dictionary zum Zwischenspeichern der Ergebnisse für den Excel-Export initialisieren
    calculated_results = {}
    
    # Liste für die CSV-Ausgabe initialisieren
    export_data = []
    
    for node_id, elements_dict in list(node_connections.items())[start_index:end_index]:
        total_elements = len(elements_dict)
        
        # Logik für Bezeichnung der Anschlussvariante (Kategoriename)
        if total_elements <= 1:
            category_name = "..."
        else:
            # Zählt die Vorkommen der Typen für diesen Knoten
            type_counts = {t: 0 for t in type_order}
            for elem_type in elements_dict.values():
                type_counts[elem_type] += 1
            
            # Baut den String in der strikten Reihenfolge auf
            name_parts = []
            for t in type_order:
                if type_counts[t] > 0:
                    name_parts.append(f"{t}{type_counts[t]}")
            
            category_name = "_".join(name_parts)

        element_names = list(elements_dict.keys())
        elements_string = ', '.join(element_names) if element_names else 'Keine'

        # Konsolenausgabe - detailliert
        # print(f"Knoten {node_id}:")
        # print(f"  - Kategorie: {category_name}")
        # print(f"  - Beteiligte Elemente ({total_elements}): {elements_string}")
        # print("-" * 30)

        # Konsolenausgabe - kompakt
        # print(f"Knoten {node_id}: {category_name}")
        # print(f"{category_name}")

        # Daten für die spätere Excel-Ausgabe zwischenspeichern
        calculated_results[node_id] = {
            'Anschlussvariante': category_name,
            'Anzahl': total_elements,
            'Elemente': elements_string
        }
        
        # Datenzeile für CSV-Format anfügen
        export_data.append({
            'Knoten': node_id,
            'Anschlussvariante': category_name,
            'Anzahl': total_elements,
            'Elemente': elements_string
        })
    # ----------------------------------------------------

    
    # --- Schritt 3: Export der Ergebnisse ---
    # Ergebnisse in bestehende Excel-Datei schreiben
    print("\nSchreibe Ergebnisse in die Excel-Datei zurück...")
    if export_excel:
        try:
            # Arbeitsmappe laden
            wb = openpyxl.load_workbook(path_saf)

            # Tabellenblatt dynamisch finden
            sheet_name = 'StructuralPointConnection'
            if sheet_name not in wb.sheetnames:
                print(f"FEHLER: Tabellenblatt '{sheet_name}' in der Excel-Datei nicht gefunden.")
                return
        
            ws = wb[sheet_name]

            # Suchen der Spalte "Name" und der ersten leeren Spalte
            name_col_idx = None
            max_col_idx = ws.max_column

            for col in range(1, max_col_idx + 1):
                cell_value = ws.cell(row=1, column=col).value
                if cell_value and str(cell_value).strip() == 'Name':
                    name_col_idx = col
                    break
            
            if not name_col_idx:
                print("FEHLER: Spalte 'Name' nicht gefunden.")
                return
            
            # Indizes für die neuen Spalten bestimmen
            label_col_idx = max_col_idx + 1
            cat_col_idx = max_col_idx + 2
            count_col_idx = max_col_idx + 3
            elem_col_idx = max_col_idx + 4
            ergänzt_col_idx = max_col_idx + 5
            entfernt_col_idx = max_col_idx + 6
            kommentar_col_idx = max_col_idx + 7

            # Header schreiben
            ws.cell(row=1, column=label_col_idx, value='Label')
            ws.cell(row=1, column=cat_col_idx, value='Anschlussvariante')
            ws.cell(row=1, column=count_col_idx, value='Anzahl')
            ws.cell(row=1, column=elem_col_idx, value='Elemente')
            ws.cell(row=1, column=ergänzt_col_idx, value='fehlend')
            ws.cell(row=1, column=entfernt_col_idx, value='zuviel')
            ws.cell(row=1, column=kommentar_col_idx, value='Kommentar')

            # Zeile für Zeile durchlaufen und Werte aus calculated_results eintragen
            for row in range(2, ws.max_row + 1):
                node_name = ws.cell(row=row, column=name_col_idx).value
                if node_name and str(node_name).strip() in calculated_results:
                    result = calculated_results[str(node_name)]
                    ws.cell(row=row, column=cat_col_idx, value=result['Anschlussvariante'])
                    ws.cell(row=row, column=count_col_idx, value=result['Anzahl'])
                    ws.cell(row=row, column=elem_col_idx, value=result['Elemente'])
            
            # --- Bedingte Formatierung hinzufügen ---
            # 1. Spalten dynamisch suchen
            first_col_idx = 1
            last_col_idx = kommentar_col_idx

            first_col_letter = get_column_letter(first_col_idx)
            label_col_letter = get_column_letter(label_col_idx)
            last_col_letter = get_column_letter(last_col_idx)
            data_range = f"${first_col_letter}1:${last_col_letter}{ws.max_row}"

            # 2. Farb-Hex-Codes (entsprechend Excel-Standard-Design)
            fill_x = PatternFill(start_color='C6E0B4', end_color='C6E0B4', fill_type='solid') # Olivgrün, Akzent 3, heller 60%
            fill_o = PatternFill(start_color='F8CBAD', end_color='F8CBAD', fill_type='solid') # Orange, Akzent 6, heller 60%
            fill_dot = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid') # Weiß, Hintergrund 1, dunkler 15%

            # 3. Formeln für die bedingte Formatierung anlegen
            rule_x = FormulaRule(formula=[f'${label_col_letter}1="x"'], fill=fill_x)
            rule_o = FormulaRule(formula=[f'${label_col_letter}1="o"'], fill=fill_o)
            rule_dot = FormulaRule(formula=[f'${label_col_letter}1="."'], fill=fill_dot)

            # 4. Regeln an Excel übergeben
            ws.conditional_formatting.add(data_range, rule_x)
            ws.conditional_formatting.add(data_range, rule_o)
            ws.conditional_formatting.add(data_range, rule_dot)
            print(">>> Bedingte Formatierungen (x, o, .) wurden erfolgreich in Excel integriert! <<<")
            # ---------------------------------------------
            
            # Datei abspeichern
            wb.save(path_saf)
            print(f">>> ERFOLG: Ergebnisse wurden in '{path_saf}' zurückgeschrieben.")

        except PermissionError:
            print(f"\n!!! KRITISCHER FEHLER: Die Datei '{path_saf}' ist in Excel geöffnet! Bitte schließe die Datei und starte das Skript erneut. !!!")
        except Exception as e:
            print(f"\n!!! FEHLER beim Schreiben in die Excel-Datei: {e} !!!")

    # Ergbnisse als DataFrame aus der Liste erstellen und als CSV speichern
    if export_csv:
        print("\nSchreibe Ergebnisse in CSV-Datei...")
        if export_data:        
            df_export = pd.DataFrame(export_data)
            df_export.to_csv(output_csv, index=False, 
                                sep=';', encoding='utf-8-sig')
            print(f">>> ERFOLG: Die Ergebnisse wurden in '{output_csv}' gespeichert. <<<")

    # Ausgabe der Warnung am Ende des Skripts
    if unrecognized_types:
        print("\n!!! ANALYSE: Unbekannte Bauteiltypen gefunden !!!")
        print("Folgende Werte konnten nicht zugeordnet werden:")
        print(", ".join(unrecognized_types))
        
    if missing_nodes:
        print("\n!!! WARNUNG: Inkonsistente Knotenreferenzen !!!")
        print(", ".join(sorted(missing_nodes)))





"""----- ZWEITER TEIL ----- ZWEITER TEIL ----- ZWEITER TEIL -----"""





def generate_gnn_labels(model_stem, export_csv=True, export_json=True):
    # --- Logger-Funktion einrichten ---
    logs = []
    def log_msg(msg, console_only=False):
        """Speichert eine Nachricht als Protkoll und gibt sie in der Konsole aus"""
        print(msg)
        if not console_only:
            # Splittet mehrzeilige Strings sauber auf, damit sie zeilenweise exportiert werden.
            for line in str(msg).split('\n'):
                clean_msg = line.strip()
                if clean_msg:
                    logs.append(clean_msg)
        return                

    log_msg(f"Lese Datei ein: {model_stem}...\n")
    # --- Schritt 0: Dateipfade vorbereiten ---
    # Basis-Pfad-Definition
    base_dir = os.path.dirname(__file__)
    saf_dir = os.path.join(base_dir, "Analysemodelle")
    dir_2_3 = os.path.join(base_dir, "2_3 Edge_Labeling")

    os.makedirs(dir_2_3, exist_ok=True) # Zielordner erstellen. falls er nicht existiert

    # Dateinamen anpassen, letzte Bezeichnung ("0D Ausr Anschluss") kürzen
    cut_str = str("Ausr Anschluss")
    if model_stem.endswith(cut_str):
        # Entfernt letzte Zeichen + 3 Zeichen davor (z.B. " 0D")
        cut_length = len(cut_str) + 3
        ifc_name = model_stem[:-cut_length].strip()
    else:
        ifc_name = model_stem.strip()

    # Dateipfade zusammenbauen
    path_saf = os.path.join(saf_dir, f"{model_stem}.xlsx")
    output_json = os.path.join(dir_2_3, f"{ifc_name} Labels.json")
    output_csv = os.path.join(dir_2_3, f"{ifc_name} Labels.csv")
    # ----------------------------------------------------


    # --- Schritt 1: Mapping und Total Labels generieren ---
    try:
        xls = pd.ExcelFile(path_saf)
    except Exception as e:
        log_msg(f"Kritischer Fehler beim Öffnen der Excel-Datei: {e}")
        return
    
    sheet_names = xls.sheet_names

    # Standardmäßig erzeugte Blätter prüfen
    required_sheets = ['StructuralPointConnection', 'StructuralCurveMember', 'StructuralSurfaceMember']
    for sheet in required_sheets:
        if sheet not in sheet_names:
            log_msg(f"KRITISCHER FEHLER: Das zwingend erforderliche Blatt '{sheet}' fehlt in der Datei.")
            return
    # Zwingende Blätter einlesen
    # Als Input geht sowohl xls als auch path_saf --> Zur Demonstration habe ich beides drin
    df_nodes = pd.read_excel(xls, sheet_name='StructuralPointConnection')
    df_curve = pd.read_excel(xls, sheet_name='StructuralCurveMember')
    df_surface = pd.read_excel(xls, sheet_name='StructuralSurfaceMember')

    # Optionale Blätter einlesen (mit Fallback auf leere Blätter)
    if 'StructuralCurveMemberRib' in sheet_names:
        df_rib = pd.read_excel(path_saf, sheet_name='StructuralCurveMemberRib')
    else:
        log_msg(f"[INFO] Tabellenblatt 'StructuralCurveMemberRib' fehlt. Rippen werden ignoriert.")
        df_rib = pd.DataFrame(columns=['Name', 'Type', 'Id', '2D member', 'Nodes'])
    
    # Spalten bereinigen
    for df in [df_nodes, df_curve, df_rib, df_surface]:
        df.columns = df.columns.str.strip()

    saf_to_ifc = {}
    all_ifc_guids = set()

    # SAF und IFC-Elemente mappen
    for df in [df_curve, df_rib, df_surface]:
        # Prüfen ob GUID existiert
        if 'GUID' not in df.columns:
            continue

        for _, row in df.iterrows():
            saf_name = str(row.get('Name')).strip()
            ifc_guid = str(row.get('GUID')).strip()

            if saf_name != 'nan' and ifc_guid != 'nan':
                saf_to_ifc[saf_name] = ifc_guid
                all_ifc_guids.add(ifc_guid)
    
    log_msg(f"Gefundene einzigartige IFC-Bauteile: {len(all_ifc_guids)}")

    # Total Labels: Alle möglichen Kombinationen im ganzen Bauwerk (ungerichtet)
    Total_Labels = set(tuple(sorted(comb)) for comb in itertools.combinations(all_ifc_guids, 2))
    log_msg(f"Generierte Total_Labels (Alle möglichen Kanten): {len(Total_Labels)}")
    # ----------------------------------------------------


    # --- SCHRITT 2: GroundTruth (GT) und Rule_Labels ---
    GT_Labels = set()
    Rule_Labels = set()

    # Hilfskunktion zum Verarbeiten von Strings wie "CM25-SFC46, CM25-SFC47"
    def parse_edges(edge_string):
        parsed_edges = set()
        if pd.isna(edge_string) or str(edge_string).strip() == '':
            return parsed_edges
        
        parts = str(edge_string).split(',')
        for part in parts:
            if '-' in part:
                elements = part.split('-', 1)
                if len(elements) == 2:
                    saf_a, saf_b = elements[0].strip(), elements[1].strip()
                    # Nur hinzufügen, wenn beide SAF-Namen eine GUID haben
                    if saf_a in saf_to_ifc and saf_b in saf_to_ifc:
                        ifc_a = saf_to_ifc[saf_a]
                        ifc_b = saf_to_ifc[saf_b]
                        edge_tuple = tuple(sorted((ifc_a, ifc_b)))
                        parsed_edges.add(edge_tuple)
        return parsed_edges
        
    
    for _, row in df_nodes.iterrows():
        label = str(row.get('Label')).strip().lower()
        if label not in ['x', 'o']:
            continue

        elemente_str = str(row.get('Elemente'))

        # GT_Labels für diesen Knoten generieren
        saf_elements_at_node = [e.strip() for e in elemente_str.split(',')]
        ifc_elements_at_node = set()

        for saf in saf_elements_at_node:
            if saf in saf_to_ifc:
                ifc_elements_at_node.add(saf_to_ifc[saf])

        # Alle Kombinationen der beteiligten Elemente am Knoten = Ground Truth
        node_gt_edges = set(tuple(sorted(comb)) for comb in itertools.combinations(ifc_elements_at_node, 2))
        GT_Labels.update(node_gt_edges)

        # Modus-Logik für Rule_Labels
        node_rule_edges = set(node_gt_edges)  # Startet mit GT-Labels als perfekte Basis

        fehlend_str = str(row.get('fehlend')).strip().lower()
        zuviel_str = str(row.get('zuviel')).strip().lower()

        # 1. Fehlende Verbindungen abziehen
        if fehlend_str == 'alle':
            # Konverter hat hier keine Elemente miteinander verbunden, trotz Vollkopplung.
            node_rule_edges.clear() 
        elif fehlend_str != 'nan':
            edges_to_remove = parse_edges(row.get('fehlend'))
            node_rule_edges.difference_update(edges_to_remove)
        
        # 2. Übermäßige Verbindungen hinzufügen
        if zuviel_str != 'nan':
            edges_to_add = parse_edges(row.get('zuviel'))
            node_rule_edges.update(edges_to_add)
        
        Rule_Labels.update(node_rule_edges)
    
    log_msg(f"Generierte GT_Labels: {len(GT_Labels)}")
    log_msg(f"Generierte Rule_Labels: {len(Rule_Labels)}")
    # ----------------------------------------------------


    # --- SCHRITT 3: Metriken berechnen ---
    tp = len(GT_Labels.intersection(Rule_Labels)) # Prüft die Schnittmenge beider Labels
    fp = len(Rule_Labels.difference(GT_Labels)) # Realität sagt "ja", Konverter hat sie übersehen
    fn = len(GT_Labels.difference(Rule_Labels)) # Realität sagt "nein", Konverter hat sie fälschlicherweise verbunden
    tn = len(Total_Labels.difference(GT_Labels.union(Rule_Labels))) # Keine Verbindung in Realität und Konverter

    total_possible = len(Total_Labels)
    accuracy = (tp + tn) / total_possible if total_possible > 0 else 0
    precision = tp / (tp + fp) if (tp + fp) > 0 else 0
    recall = tp / (tp + fn) if (tp + fn) > 0 else 0
    f1_score = 2 * (precision * recall) / (precision + recall) if (precision + recall) > 0 else 0

    log_msg("\n=== AUSWERTUNG DER KONVERTER-REGELN (Metriken) ===")
    log_msg(f"True Positives (TP):  {tp:<6} (Richtig erkannt)")
    log_msg(f"True Negatives (TN):  {tn:<6} (Richtig ignoriert)")
    log_msg(f"False Positives (FP): {fp:<6} (Zu viel verbunden)")
    log_msg(f"False Negatives (FN): {fn:<6} (Vergessen)")
    log_msg("-" * 50)
    log_msg(f"Accuracy:  {accuracy:.4f} ({(accuracy*100):.2f} %)")
    log_msg(f"Precision: {precision:.4f} ({(precision*100):.2f} %)")
    log_msg(f"Recall:    {recall:.4f} ({(recall*100):.2f} %)")
    log_msg(f"F1-Score:  {f1_score:.4f}")
    log_msg("==================================================")
    # ----------------------------------------------------


    # --- SCHRITT 4: Ergebnisse dynamisch exportieren ---
    # Konvertierung in strukturierte Dictionaries
    def set_to_dict_list(edge_set, label_type):
        return [{'GUID_A': edge[0], 'GUID_B': edge[1], 'Label_Type': label_type} for edge in edge_set]
    
    edge_labels_list = set_to_dict_list(Total_Labels, 'Total') + \
                set_to_dict_list(GT_Labels, 'GT') + \
                set_to_dict_list(Rule_Labels, 'Rule')
    
    json_export_data = {
        "stats_Rule_conversion": {
            "IFC elements": len(all_ifc_guids),
            "total_edges": total_possible,
            "count_GT_labels": len(GT_Labels),
            "count_Rule_labels": len(Rule_Labels),
            "TP": tp,
            "TN": tn,
            "FP": fp,
            "FN": fn,
            "Accuracy": round(accuracy, 4),
            "Precision": round(precision, 4),
            "Recall": round(recall, 4),
            "F1_Score": round(f1_score, 4)
        },
        "edge_labels": edge_labels_list
    }
    
    # --- JSON-Export ---
    if export_json:
        with open(output_json, 'w', encoding='utf-8') as f:
            json.dump(json_export_data, f, indent=4)
    log_msg(f"JSON gespeichert: {output_json}", console_only=True)
    
    # --- CSV-Export ---
    df_export = pd.DataFrame(edge_labels_list)

    # Sicherstellen, dass die Tabelle lang genug ist, um gesamtes Log aufzunehmen
    max_len = max(len(df_export), len(logs))
    if len(df_export) < max_len:
        empty_rows = pd.DataFrame([{}] * (max_len - len(df_export)))
        df_export = pd.concat([df_export, empty_rows], ignore_index=True)

    logs_padded = logs + [""] * (max_len - len(logs))

    # 2 Spalten freilassen (über Leerzeichen im Header differenziert)
    df_export[' '] = ""   # Leere Spalte 1
    df_export['  '] = ""  # Leere Spalte 2
    df_export['Konsolen-Protokoll'] = logs_padded
    # --------------------------------------------------------

    if export_csv:
        df_export.to_csv(output_csv, index=False, sep=';', encoding='utf-8-sig')
        log_msg(f"CSV gespeichert:  {output_csv}", console_only=True)

    print(f"\n>>> ERFOLG: GNN-Labels wurden in '{output_csv}' und '{output_json}' gespeichert. <<<")
    # print(f"\n--- ___KOPIERE DEN TEXT IN DIE CSV !!___ ---")
    # ----------------------------------------------------





"""Main-Part"""
if __name__ == "__main__":
    main()
